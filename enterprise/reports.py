"""
Enterprise Reports - 企业报表

提供企业级报表功能：
- EnterpriseReport: 企业报表模型
- ReportGenerator: 报表生成器

Version: v0.9.0
"""

from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from enum import Enum
import json


class ReportType(str, Enum):
    """报表类型"""
    TENANT_SUMMARY = "tenant_summary"
    TEAM_PERFORMANCE = "team_performance"
    USER_ACTIVITY = "user_activity"
    SESSION_ANALYSIS = "session_analysis"
    PROGRESS_TRACKING = "progress_tracking"
    EXECUTIVE_DASHBOARD = "executive_dashboard"


class ReportFormat(str, Enum):
    """报表格式"""
    JSON = "json"
    PDF = "pdf"
    EXCEL = "excel"
    HTML = "html"


@dataclass
class TeamReport:
    """团队报表"""
    team_id: str
    period_days: int = 30
    sections: List[str] = field(default_factory=list)
    data: Dict[str, Any] = field(default_factory=dict)
    generated_at: datetime = field(default_factory=datetime.utcnow)


@dataclass
class EnterpriseReport:
    """企业报表"""
    id: str
    report_type: ReportType
    tenant_id: str
    title: str
    description: Optional[str] = None
    format: ReportFormat = ReportFormat.JSON
    status: str = "pending"  # pending, completed, failed
    created_at: datetime = field(default_factory=datetime.utcnow)
    completed_at: Optional[datetime] = None
    period_start: Optional[datetime] = None
    period_end: Optional[datetime] = None
    filters: Dict[str, Any] = field(default_factory=dict)
    data: Dict[str, Any] = field(default_factory=dict)
    summary: Dict[str, Any] = field(default_factory=dict)
    file_path: Optional[str] = None
    file_size_bytes: int = 0
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self):
        if isinstance(self.report_type, str):
            self.report_type = ReportType(self.report_type)
        if isinstance(self.format, str):
            self.format = ReportFormat(self.format)

    @property
    def is_completed(self) -> bool:
        """检查报表是否已完成"""
        return self.status == "completed"

    @property
    def is_failed(self) -> bool:
        """检查报表是否失败"""
        return self.status == "failed"

    @property
    def generation_time_seconds(self) -> Optional[float]:
        """计算生成时间（秒）"""
        if self.completed_at:
            return (self.completed_at - self.created_at).total_seconds()
        return None

    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return {
            "id": self.id,
            "report_type": self.report_type.value,
            "tenant_id": self.tenant_id,
            "title": self.title,
            "description": self.description,
            "format": self.format.value,
            "status": self.status,
            "created_at": self.created_at.isoformat(),
            "completed_at": self.completed_at.isoformat() if self.completed_at else None,
            "period_start": self.period_start.isoformat() if self.period_start else None,
            "period_end": self.period_end.isoformat() if self.period_end else None,
            "filters": self.filters,
            "data": self.data,
            "summary": self.summary,
            "file_path": self.file_path,
            "file_size_bytes": self.file_size_bytes,
            "metadata": self.metadata,
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "EnterpriseReport":
        """从字典创建"""
        if "created_at" in data and isinstance(data["created_at"], str):
            data["created_at"] = datetime.fromisoformat(data["created_at"])
        if "completed_at" in data and isinstance(data["completed_at"], str):
            data["completed_at"] = datetime.fromisoformat(data["completed_at"])
        if "period_start" in data and isinstance(data["period_start"], str):
            data["period_start"] = datetime.fromisoformat(data["period_start"])
        if "period_end" in data and isinstance(data["period_end"], str):
            data["period_end"] = datetime.fromisoformat(data["period_end"])
        return cls(**data)


class ReportGenerator:
    """报表生成器"""

    def __init__(
        self,
        tenant_manager=None,
        team_manager=None,
        collaboration_manager=None,
        progress_tracker=None,
    ):
        self.tenant_manager = tenant_manager
        self.team_manager = team_manager
        self.collaboration_manager = collaboration_manager
        self.progress_tracker = progress_tracker
        self._reports: Dict[str, EnterpriseReport] = {}

    def generate_report(
        self,
        report_type: ReportType,
        tenant_id: str,
        title: Optional[str] = None,
        period_days: int = 30,
        filters: Optional[Dict[str, Any]] = None,
        output_format: ReportFormat = ReportFormat.JSON,
    ) -> EnterpriseReport:
        """生成报表"""
        import hashlib
        report_id = f"report_{hashlib.md5(f'{tenant_id}{report_type}{datetime.utcnow().isoformat()}'.encode()).hexdigest()[:12]}"

        period_end = datetime.utcnow()
        period_start = period_end - timedelta(days=period_days)

        report = EnterpriseReport(
            id=report_id,
            report_type=report_type,
            tenant_id=tenant_id,
            title=title or self._get_default_title(report_type),
            format=output_format,
            period_start=period_start,
            period_end=period_end,
            filters=filters or {},
        )

        self._reports[report_id] = report

        # 异步生成报表数据
        try:
            report.data = self._generate_report_data(report_type, tenant_id, period_start, period_end, filters)
            report.summary = self._generate_summary(report_type, report.data)
            report.status = "completed"
            report.completed_at = datetime.utcnow()
        except Exception as e:
            report.status = "failed"
            report.metadata["error"] = str(e)

        return report

    def generate_team_report(
        self,
        team_id: str,
        period_days: int = 30,
        **kwargs,
    ) -> "TeamReport":
        """生成团队报表"""
        # Generate a simple team report
        return TeamReport(
            team_id=team_id,
            period_days=period_days,
            sections=["overview", "performance", "activity", "recommendations"],
        )

    def get_report(self, report_id: str) -> Optional[EnterpriseReport]:
        """获取报表"""
        return self._reports.get(report_id)

    def list_reports(
        self,
        tenant_id: Optional[str] = None,
        report_type: Optional[ReportType] = None,
        status: Optional[str] = None,
    ) -> List[EnterpriseReport]:
        """列出报表"""
        reports = list(self._reports.values())

        if tenant_id:
            reports = [r for r in reports if r.tenant_id == tenant_id]
        if report_type:
            reports = [r for r in reports if r.report_type == report_type]
        if status:
            reports = [r for r in reports if r.status == status]

        return sorted(reports, key=lambda r: r.created_at, reverse=True)

    def export_report(self, report_id: str, format: Optional[ReportFormat] = None) -> Optional[str]:
        """导出报表"""
        report = self._reports.get(report_id)
        if not report:
            return None

        export_format = format or report.format

        if export_format == ReportFormat.JSON:
            return self._export_json(report)
        elif export_format == ReportFormat.EXCEL:
            return self._export_excel(report)
        elif export_format == ReportFormat.HTML:
            return self._export_html(report)
        elif export_format == ReportFormat.PDF:
            return self._export_pdf(report)

        return None

    def _get_default_title(self, report_type: ReportType) -> str:
        """获取默认标题"""
        titles = {
            ReportType.TENANT_SUMMARY: "Tenant Summary Report",
            ReportType.TEAM_PERFORMANCE: "Team Performance Report",
            ReportType.USER_ACTIVITY: "User Activity Report",
            ReportType.SESSION_ANALYSIS: "Session Analysis Report",
            ReportType.PROGRESS_TRACKING: "Progress Tracking Report",
            ReportType.EXECUTIVE_DASHBOARD: "Executive Dashboard",
        }
        return titles.get(report_type, "Enterprise Report")

    def _generate_report_data(
        self,
        report_type: ReportType,
        tenant_id: str,
        period_start: datetime,
        period_end: datetime,
        filters: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """生成报表数据"""
        if report_type == ReportType.TENANT_SUMMARY:
            return self._generate_tenant_summary(tenant_id, period_start, period_end)
        elif report_type == ReportType.TEAM_PERFORMANCE:
            return self._generate_team_performance(tenant_id, period_start, period_end, filters)
        elif report_type == ReportType.USER_ACTIVITY:
            return self._generate_user_activity(tenant_id, period_start, period_end)
        elif report_type == ReportType.SESSION_ANALYSIS:
            return self._generate_session_analysis(tenant_id, period_start, period_end, filters)
        elif report_type == ReportType.PROGRESS_TRACKING:
            return self._generate_progress_tracking(tenant_id, period_start, period_end)
        elif report_type == ReportType.EXECUTIVE_DASHBOARD:
            return self._generate_executive_dashboard(tenant_id, period_start, period_end)
        else:
            return {}

    def _generate_tenant_summary(
        self,
        tenant_id: str,
        period_start: datetime,
        period_end: datetime,
    ) -> Dict[str, Any]:
        """生成租户摘要"""
        if not self.tenant_manager:
            return {}

        tenant = self.tenant_manager.get_tenant(tenant_id)
        if not tenant:
            return {}

        stats = self.tenant_manager.get_statistics()

        return {
            "tenant_info": tenant.to_dict(),
            "statistics": stats,
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
        }

    def _generate_team_performance(
        self,
        tenant_id: str,
        period_start: datetime,
        period_end: datetime,
        filters: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """生成团队表现报表"""
        if not self.team_manager:
            return {}

        teams = self.team_manager.list_teams(tenant_id=tenant_id)
        team_data = []

        for team in teams:
            team_stats = self.team_manager.get_statistics(team.tenant_id)

            collaboration_stats = {}
            if self.collaboration_manager:
                collaboration_stats = self.collaboration_manager.get_statistics(team.id)

            team_data.append({
                "team_id": team.id,
                "team_name": team.name,
                "member_count": team.member_count,
                "status": team.status.value,
                "statistics": team_stats,
                "collaboration": collaboration_stats,
            })

        return {
            "teams": team_data,
            "total_teams": len(teams),
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
        }

    def _generate_user_activity(
        self,
        tenant_id: str,
        period_start: datetime,
        period_end: datetime,
    ) -> Dict[str, Any]:
        """生成用户活动报表"""
        # 简化实现
        return {
            "tenant_id": tenant_id,
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
            "activity_summary": {
                "total_logins": 0,
                "total_sessions": 0,
                "active_users": 0,
            },
        }

    def _generate_session_analysis(
        self,
        tenant_id: str,
        period_start: datetime,
        period_end: datetime,
        filters: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """生成会话分析报表"""
        if not self.collaboration_manager:
            return {}

        all_sessions = []
        teams = self.team_manager.list_teams(tenant_id=tenant_id) if self.team_manager else []

        for team in teams:
            sessions = self.collaboration_manager.list_sessions(team_id=team.id)
            all_sessions.extend(sessions)

        # 过滤
        filtered_sessions = [
            s for s in all_sessions
            if s.started_at and period_start <= s.started_at <= period_end
        ]

        # 按类型统计
        by_type = {}
        for session in filtered_sessions:
            type_key = session.session_type.value
            if type_key not in by_type:
                by_type[type_key] = {"count": 0, "total_duration": 0}
            by_type[type_key]["count"] += 1
            by_type[type_key]["total_duration"] += session.duration_minutes or 0

        return {
            "sessions": [s.to_dict() for s in filtered_sessions],
            "total_sessions": len(filtered_sessions),
            "by_type": by_type,
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
        }

    def _generate_progress_tracking(
        self,
        tenant_id: str,
        period_start: datetime,
        period_end: datetime,
    ) -> Dict[str, Any]:
        """生成进度追踪报表"""
        if not self.progress_tracker:
            return {}

        # 简化实现
        return {
            "tenant_id": tenant_id,
            "period": {
                "start": period_start.isoformat(),
                "end": period_end.isoformat(),
            },
            "progress_summary": {
                "total_users": 0,
                "improving_users": 0,
                "stable_users": 0,
                "declining_users": 0,
            },
        }

    def _generate_executive_dashboard(
        self,
        tenant_id: str,
        period_start: datetime,
        period_end: datetime,
    ) -> Dict[str, Any]:
        """生成执行仪表盘"""
        tenant_summary = self._generate_tenant_summary(tenant_id, period_start, period_end)
        team_performance = self._generate_team_performance(tenant_id, period_start, period_end, None)
        session_analysis = self._generate_session_analysis(tenant_id, period_start, period_end, None)

        return {
            "tenant_summary": tenant_summary,
            "team_performance": team_performance,
            "session_analysis": session_analysis,
            "kpi": {
                "total_teams": team_performance.get("total_teams", 0),
                "total_sessions": session_analysis.get("total_sessions", 0),
                "avg_session_duration": 0,
            },
        }

    def _generate_summary(self, report_type: ReportType, data: Dict[str, Any]) -> Dict[str, Any]:
        """生成摘要"""
        # 简化实现
        return {
            "generated_at": datetime.utcnow().isoformat(),
            "data_points": len(str(data)),
        }

    def _export_json(self, report: EnterpriseReport) -> str:
        """导出为 JSON"""
        import os
        os.makedirs("data/reports", exist_ok=True)
        file_path = f"data/reports/{report.id}.json"

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(report.to_dict(), f, indent=2, ensure_ascii=False)

        report.file_path = file_path
        report.file_size_bytes = os.path.getsize(file_path) if os.path.exists(file_path) else 0
        return file_path

    def _export_excel(self, report: EnterpriseReport) -> Optional[str]:
        """导出为 Excel"""
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            return None

        import os
        os.makedirs("data/reports", exist_ok=True)
        file_path = f"data/reports/{report.id}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # 写入摘要
        ws.cell(row=1, column=1, value="Report ID")
        ws.cell(row=1, column=2, value=report.id)
        ws.cell(row=2, column=1, value="Title")
        ws.cell(row=2, column=2, value=report.title)
        ws.cell(row=3, column=1, value="Type")
        ws.cell(row=3, column=2, value=report.report_type.value)
        ws.cell(row=4, column=1, value="Status")
        ws.cell(row=4, column=2, value=report.status)

        wb.save(file_path)
        report.file_path = file_path
        report.file_size_bytes = os.path.getsize(file_path) if os.path.exists(file_path) else 0
        return file_path

    def _export_html(self, report: EnterpriseReport) -> Optional[str]:
        """导出为 HTML"""
        import os
        os.makedirs("data/reports", exist_ok=True)
        file_path = f"data/reports/{report.id}.html"

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{report.title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333; }}
                .summary {{ background: #f5f5f5; padding: 15px; border-radius: 5px; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #4CAF50; color: white; }}
            </style>
        </head>
        <body>
            <h1>{report.title}</h1>
            <div class="summary">
                <p><strong>Report ID:</strong> {report.id}</p>
                <p><strong>Type:</strong> {report.report_type.value}</p>
                <p><strong>Status:</strong> {report.status}</p>
                <p><strong>Generated:</strong> {report.created_at.strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        </body>
        </html>
        """

        with open(file_path, "w", encoding="utf-8") as f:
            f.write(html)

        report.file_path = file_path
        report.file_size_bytes = os.path.getsize(file_path) if os.path.exists(file_path) else 0
        return file_path

    def _export_pdf(self, report: EnterpriseReport) -> Optional[str]:
        """导出为 PDF"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except ImportError:
            return None

        import os
        os.makedirs("data/reports", exist_ok=True)
        file_path = f"data/reports/{report.id}.pdf"

        c = canvas.Canvas(file_path, pagesize=letter)
        width, height = letter

        # 标题
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, height - 50, report.title)

        # 信息
        c.setFont("Helvetica", 12)
        y = height - 100
        c.drawString(50, y, f"Report ID: {report.id}")
        y -= 20
        c.drawString(50, y, f"Type: {report.report_type.value}")
        y -= 20
        c.drawString(50, y, f"Status: {report.status}")
        y -= 20
        c.drawString(50, y, f"Generated: {report.created_at.strftime('%Y-%m-%d %H:%M:%S')}")

        c.save()

        report.file_path = file_path
        report.file_size_bytes = os.path.getsize(file_path) if os.path.exists(file_path) else 0
        return file_path
