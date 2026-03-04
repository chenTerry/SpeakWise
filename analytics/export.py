"""
Data Export Module

数据导出模块提供：
- DataExporter: 数据导出系统
- PDF/Excel/JSON 格式导出
- 综合报告生成
- 数据备份与恢复

Version: v0.7.0
Author: AgentScope AI Interview Team
"""

import json
import os
import tempfile
from datetime import datetime
from typing import List, Dict, Any, Optional, BinaryIO
from dataclasses import dataclass, field, asdict
from enum import Enum
from pathlib import Path


class ExportFormat(str, Enum):
    """导出格式"""
    PDF = "pdf"
    EXCEL = "excel"
    JSON = "json"
    CSV = "csv"


class ReportType(str, Enum):
    """报告类型"""
    PROGRESS = "progress"  # 进度报告
    ANALYTICS = "analytics"  # 分析报告
    COMPLETE = "complete"  # 完整报告
    SUMMARY = "summary"  # 摘要报告


@dataclass
class ExportResult:
    """导出结果"""
    success: bool
    file_path: Optional[str]
    file_size_bytes: int
    format: ExportFormat
    message: str
    created_at: datetime = field(default_factory=datetime.utcnow)


@dataclass
class BackupData:
    """备份数据"""
    version: str
    user_id: str
    backed_up_at: datetime
    sessions: List[Dict[str, Any]]
    analytics: Dict[str, Any]
    metadata: Dict[str, Any]


class DataExporter:
    """
    数据导出器

    提供多种格式的数据导出功能，支持报告生成和数据备份

    Design Principles:
    - 多格式支持：PDF、Excel、JSON 等
    - 数据完整性：导出完整准确的数据
    - 用户友好：清晰的报告格式
    - 可扩展性：易于添加新格式
    """

    # 数据版本
    DATA_VERSION = "1.0"

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化数据导出器

        Args:
            config: 配置字典
        """
        self.config = config or {}
        self.output_dir = self.config.get("output_dir", "./exports")
        self._ensure_output_dir()

    def _ensure_output_dir(self):
        """确保输出目录存在"""
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)

    # =========================================================================
    # Main Export Methods
    # =========================================================================

    def export_to_pdf(self, data: Dict[str, Any],
                      filename: Optional[str] = None,
                      report_type: ReportType = ReportType.COMPLETE) -> ExportResult:
        """
        导出为 PDF

        Args:
            data: 要导出的数据
            filename: 文件名（可选）
            report_type: 报告类型

        Returns:
            ExportResult 对象
        """
        try:
            # 尝试使用 reportlab
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
                from reportlab.lib.enums import TA_CENTER, TA_LEFT
                reportlab_available = True
            except ImportError:
                reportlab_available = False

            if not reportlab_available:
                # 降级为 JSON 导出
                return self._export_pdf_fallback(data, filename, report_type)

            filename = filename or f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            file_path = os.path.join(self.output_dir, filename)

            # 创建 PDF 文档
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=0.5 * inch,
                leftMargin=0.5 * inch,
                topMargin=0.5 * inch,
                bottomMargin=0.5 * inch
            )

            elements = []
            styles = getSampleStyleSheet()

            # 标题
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=TA_CENTER
            )

            report_titles = {
                ReportType.PROGRESS: "学习进度报告",
                ReportType.ANALYTICS: "学习分析报告",
                ReportType.COMPLETE: "综合分析报告",
                ReportType.SUMMARY: "学习摘要",
            }

            elements.append(Paragraph(report_titles.get(report_type, "报告"), title_style))
            elements.append(Spacer(1, 0.2 * inch))

            # 添加内容
            elements.extend(self._build_pdf_content(data, report_type, styles))

            # 构建 PDF
            doc.build(elements)

            file_size = os.path.getsize(file_path)

            return ExportResult(
                success=True,
                file_path=file_path,
                file_size_bytes=file_size,
                format=ExportFormat.PDF,
                message="PDF 报告生成成功"
            )

        except Exception as e:
            return ExportResult(
                success=False,
                file_path=None,
                file_size_bytes=0,
                format=ExportFormat.PDF,
                message=f"PDF 导出失败：{str(e)}"
            )

    def _build_pdf_content(self, data: Dict[str, Any],
                          report_type: ReportType,
                          styles) -> List:
        """构建 PDF 内容"""
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        elements = []

        # 用户信息
        if "user_id" in data:
            elements.append(Paragraph("用户信息", styles['Heading2']))
            user_data = [
                ["用户 ID", data.get("user_id", "N/A")],
                ["生成时间", datetime.utcnow().strftime("%Y-%m-%d %H:%M")],
            ]
            if "period_days" in data:
                user_data.append(["统计周期", f"{data['period_days']}天"])

            table = Table(user_data, colWidths=[2*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.3 * inch))

        # 关键指标
        if "key_insights" in data or "summary_text" in data:
            elements.append(Paragraph("关键洞察", styles['Heading2']))
            if "summary_text" in data:
                elements.append(Paragraph(data["summary_text"], styles['Normal']))

            if "key_insights" in data:
                for insight in data["key_insights"][:5]:
                    title = insight.get("title", "")
                    desc = insight.get("description", "")
                    elements.append(Paragraph(f"• <b>{title}</b>: {desc}", styles['Normal']))

            elements.append(Spacer(1, 0.3 * inch))

        # 表现卡片
        if "performance_cards" in data:
            elements.append(Paragraph("维度表现", styles['Heading2']))

            table_data = [["维度", "当前分数", "变化", "水平"]]
            for card in data["performance_cards"][:7]:
                change_str = f"+{card['change']:.1f}" if card['change'] > 0 else f"{card['change']:.1f}"
                table_data.append([
                    card["dimension_name"],
                    f"{card['current_score']:.1f}",
                    change_str,
                    card["level"],
                ])

            table = Table(table_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.3 * inch))

        # 建议
        if "recommendations" in data:
            elements.append(Paragraph("推荐练习", styles['Heading2']))
            for i, rec in enumerate(data["recommendations"][:5], 1):
                elements.append(Paragraph(f"{i}. <b>{rec.get('title', '')}</b>", styles['Normal']))
                elements.append(Paragraph(f"   {rec.get('description', '')}", styles['Normal']))

            elements.append(Spacer(1, 0.3 * inch))

        return elements

    def _export_pdf_fallback(self, data: Dict[str, Any],
                            filename: Optional[str],
                            report_type: ReportType) -> ExportResult:
        """PDF 导出降级方案（无 reportlab 时）"""
        # 尝试使用 matplotlib 生成简单 PDF
        try:
            import matplotlib
            matplotlib.use('Agg')  # 非交互式后端
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_pdf import PdfPages

            filename = filename or f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            file_path = os.path.join(self.output_dir, filename)

            with PdfPages(file_path) as pdf:
                # 创建图表页
                fig, axes = plt.subplots(2, 2, figsize=(10, 10))

                # 维度表现图
                if "performance_cards" in data:
                    cards = data["performance_cards"][:7]
                    names = [c["dimension_name"] for c in cards]
                    scores = [c["current_score"] for c in cards]

                    axes[0, 0].barh(names, scores, color='steelblue')
                    axes[0, 0].set_xlabel('分数')
                    axes[0, 0].set_title('维度表现')
                    axes[0, 0].invert_yaxis()

                # 趋势图
                if "trend_analyses" in data:
                    trends = data["trend_analyses"]
                    if trends:
                        names = [t["metric_name"] for t in trends]
                        values = [t["current_value"] for t in trends]
                        axes[0, 1].bar(names, values, color='coral')
                        axes[0, 1].set_ylabel('分数')
                        axes[0, 1].set_title('趋势分析')
                        axes[0, 1].tick_params(axis='x', rotation=45)

                # 成就
                if "achievements" in data:
                    achievements = data["achievements"]
                    count = len(achievements)
                    axes[1, 0].pie([count, max(0, 10-count)], labels=['已获得', '未获得'],
                                  autopct='%1.1f%%', colors=['gold', 'lightgray'])
                    axes[1, 0].set_title(f'成就 ({count}个)')

                # 文本摘要
                axes[1, 1].axis('off')
                summary = data.get("summary_text", "暂无摘要")
                axes[1, 1].text(0.1, 0.9, f"摘要:\n{summary}",
                               transform=axes[1, 1].transAxes,
                               fontsize=10, verticalalignment='top',
                               bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))

                plt.tight_layout()
                pdf.savefig(fig)
                plt.close()

            file_size = os.path.getsize(file_path)

            return ExportResult(
                success=True,
                file_path=file_path,
                file_size_bytes=file_size,
                format=ExportFormat.PDF,
                message="PDF 报告生成成功（图表版）"
            )

        except Exception as e:
            return ExportResult(
                success=False,
                file_path=None,
                file_size_bytes=0,
                format=ExportFormat.PDF,
                message=f"PDF 导出失败：{str(e)}"
            )

    def export_to_excel(self, data: Dict[str, Any],
                       filename: Optional[str] = None) -> ExportResult:
        """
        导出为 Excel

        Args:
            data: 要导出的数据
            filename: 文件名（可选）

        Returns:
            ExportResult 对象
        """
        try:
            # 尝试使用 openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                openpyxl_available = True
            except ImportError:
                openpyxl_available = False

            if not openpyxl_available:
                return ExportResult(
                    success=False,
                    file_path=None,
                    file_size_bytes=0,
                    format=ExportFormat.EXCEL,
                    message="Excel 导出需要安装 openpyxl: pip install openpyxl"
                )

            filename = filename or f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(self.output_dir, filename)

            # 创建工作簿
            wb = Workbook()

            # 删除默认 sheet
            default_sheet = wb.active
            wb.remove(default_sheet)

            # 创建摘要 sheet
            summary_sheet = wb.create_sheet("摘要")
            self._build_excel_summary_sheet(summary_sheet, data)

            # 创建表现 sheet
            if "performance_cards" in data:
                perf_sheet = wb.create_sheet("维度表现")
                self._build_excel_performance_sheet(perf_sheet, data["performance_cards"])

            # 创建洞察 sheet
            if "key_insights" in data:
                insights_sheet = wb.create_sheet("关键洞察")
                self._build_excel_insights_sheet(insights_sheet, data["key_insights"])

            # 创建建议 sheet
            if "recommendations" in data:
                rec_sheet = wb.create_sheet("推荐练习")
                self._build_excel_recommendations_sheet(rec_sheet, data["recommendations"])

            # 保存
            wb.save(file_path)
            file_size = os.path.getsize(file_path)

            return ExportResult(
                success=True,
                file_path=file_path,
                file_size_bytes=file_size,
                format=ExportFormat.EXCEL,
                message="Excel 报告生成成功"
            )

        except Exception as e:
            return ExportResult(
                success=False,
                file_path=None,
                file_size_bytes=0,
                format=ExportFormat.EXCEL,
                message=f"Excel 导出失败：{str(e)}"
            )

    def _build_excel_summary_sheet(self, sheet, data: Dict[str, Any]):
        """构建 Excel 摘要 sheet"""
        from openpyxl.styles import Font, Alignment, PatternFill

        # 标题
        sheet['A1'] = '学习分析报告'
        sheet['A1'].font = Font(bold=True, size=16)

        # 基本信息
        row = 3
        sheet[f'A{row}'] = '用户 ID'
        sheet[f'B{row}'] = data.get('user_id', 'N/A')
        row += 1

        sheet[f'A{row}'] = '生成时间'
        sheet[f'B{row}'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        row += 1

        if 'period_days' in data:
            sheet[f'A{row}'] = '统计周期'
            sheet[f'B{row}'] = f"{data['period_days']} 天"
            row += 1

        # 摘要文本
        row += 1
        sheet[f'A{row}'] = '摘要'
        sheet[f'A{row}'].font = Font(bold=True)
        row += 1
        sheet[f'A{row}'] = data.get('summary_text', '')

        # 关键指标
        row += 2
        sheet[f'A{row}'] = '关键指标'
        sheet[f'A{row}'].font = Font(bold=True)

    def _build_excel_performance_sheet(self, sheet, performance_cards: List[Dict]):
        """构建 Excel 表现 sheet"""
        from openpyxl.styles import Font, PatternFill

        # 表头
        headers = ['维度', '当前分数', '之前分数', '变化', '百分位', '水平', '状态']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')

        # 数据
        for row, card in enumerate(performance_cards, 2):
            sheet.cell(row=row, column=1, value=card.get('dimension_name', ''))
            sheet.cell(row=row, column=2, value=card.get('current_score', 0))
            sheet.cell(row=row, column=3, value=card.get('previous_score', 0))
            change = card.get('change', 0)
            sheet.cell(row=row, column=4, value=f"+{change:.1f}" if change > 0 else f"{change:.1f}")
            sheet.cell(row=row, column=5, value=f"{card.get('percentile', 0):.0f}%")
            sheet.cell(row=row, column=6, value=card.get('level', ''))
            sheet.cell(row=row, column=7, value=card.get('status', ''))

        # 调整列宽
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column].width = min(max_length + 2, 20)

    def _build_excel_insights_sheet(self, sheet, insights: List[Dict]):
        """构建 Excel 洞察 sheet"""
        # 表头
        headers = ['类别', '标题', '描述', '趋势', '行动建议']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # 数据
        for row, insight in enumerate(insights, 2):
            sheet.cell(row=row, column=1, value=insight.get('category', ''))
            sheet.cell(row=row, column=2, value=insight.get('title', ''))
            sheet.cell(row=row, column=3, value=insight.get('description', ''))
            sheet.cell(row=row, column=4, value=insight.get('trend', ''))
            actions = ', '.join(insight.get('actionable_items', []))
            sheet.cell(row=row, column=5, value=actions)

    def _build_excel_recommendations_sheet(self, sheet, recommendations: List[Dict]):
        """构建 Excel 建议 sheet"""
        # 表头
        headers = ['优先级', '练习名称', '描述', '预计影响', '所需时间 (分钟)']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # 数据
        for row, rec in enumerate(recommendations, 2):
            sheet.cell(row=row, column=1, value=rec.get('priority', 0))
            sheet.cell(row=row, column=2, value=rec.get('title', ''))
            sheet.cell(row=row, column=3, value=rec.get('description', ''))
            sheet.cell(row=row, column=4, value=rec.get('estimated_impact', ''))
            sheet.cell(row=row, column=5, value=rec.get('time_required_minutes', 0))

    def export_to_json(self, data: Dict[str, Any],
                      filename: Optional[str] = None,
                      pretty: bool = True) -> ExportResult:
        """
        导出为 JSON

        Args:
            data: 要导出的数据
            filename: 文件名（可选）
            pretty: 是否格式化输出

        Returns:
            ExportResult 对象
        """
        try:
            filename = filename or f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
            file_path = os.path.join(self.output_dir, filename)

            # 添加元数据
            export_data = {
                "export_metadata": {
                    "exported_at": datetime.utcnow().isoformat(),
                    "version": self.DATA_VERSION,
                    "format": ExportFormat.JSON.value,
                },
                "data": data,
            }

            with open(file_path, 'w', encoding='utf-8') as f:
                if pretty:
                    json.dump(export_data, f, ensure_ascii=False, indent=2)
                else:
                    json.dump(export_data, f, ensure_ascii=False)

            file_size = os.path.getsize(file_path)

            return ExportResult(
                success=True,
                file_path=file_path,
                file_size_bytes=file_size,
                format=ExportFormat.JSON,
                message="JSON 导出成功"
            )

        except Exception as e:
            return ExportResult(
                success=False,
                file_path=None,
                file_size_bytes=0,
                format=ExportFormat.JSON,
                message=f"JSON 导出失败：{str(e)}"
            )

    def export_to_csv(self, sessions: List[Dict[str, Any]],
                     filename: Optional[str] = None) -> ExportResult:
        """
        导出会话数据为 CSV

        Args:
            sessions: 会话数据列表
            filename: 文件名（可选）

        Returns:
            ExportResult 对象
        """
        try:
            import csv

            filename = filename or f"sessions_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            file_path = os.path.join(self.output_dir, filename)

            if not sessions:
                return ExportResult(
                    success=False,
                    file_path=None,
                    file_size_bytes=0,
                    format=ExportFormat.CSV,
                    message="没有会话数据可导出"
                )

            # CSV 表头
            fieldnames = [
                'session_id', 'completed_at', 'scene_type', 'duration_seconds',
                'overall_quality', 'language_expression', 'logical_thinking',
                'professional_knowledge', 'problem_solving',
                'communication_collaboration', 'adaptability'
            ]

            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for session in sessions:
                    eval_result = session.get('evaluation_result', {})
                    row = {
                        'session_id': session.get('session_id', ''),
                        'completed_at': session.get('completed_at', ''),
                        'scene_type': session.get('scene_type', ''),
                        'duration_seconds': session.get('duration_seconds', 0),
                        'overall_quality': eval_result.get('overall_quality', 0),
                        'language_expression': eval_result.get('language_expression', 0),
                        'logical_thinking': eval_result.get('logical_thinking', 0),
                        'professional_knowledge': eval_result.get('professional_knowledge', 0),
                        'problem_solving': eval_result.get('problem_solving', 0),
                        'communication_collaboration': eval_result.get('communication_collaboration', 0),
                        'adaptability': eval_result.get('adaptability', 0),
                    }
                    writer.writerow(row)

            file_size = os.path.getsize(file_path)

            return ExportResult(
                success=True,
                file_path=file_path,
                file_size_bytes=file_size,
                format=ExportFormat.CSV,
                message="CSV 导出成功"
            )

        except Exception as e:
            return ExportResult(
                success=False,
                file_path=None,
                file_size_bytes=0,
                format=ExportFormat.CSV,
                message=f"CSV 导出失败：{str(e)}"
            )

    # =========================================================================
    # Backup & Restore
    # =========================================================================

    def create_backup(self, user_id: str,
                     sessions: List[Dict[str, Any]],
                     analytics: Optional[Dict[str, Any]] = None,
                     filename: Optional[str] = None) -> ExportResult:
        """
        创建数据备份

        Args:
            user_id: 用户 ID
            sessions: 会话数据
            analytics: 分析数据
            filename: 文件名（可选）

        Returns:
            ExportResult 对象
        """
        try:
            filename = filename or f"backup_{user_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
            file_path = os.path.join(self.output_dir, filename)

            backup_data = BackupData(
                version=self.DATA_VERSION,
                user_id=user_id,
                backed_up_at=datetime.utcnow(),
                sessions=sessions,
                analytics=analytics or {},
                metadata={
                    "session_count": len(sessions),
                    "backup_type": "full",
                }
            )

            # 转换为字典
            backup_dict = {
                "version": backup_data.version,
                "user_id": backup_data.user_id,
                "backed_up_at": backup_data.backed_up_at.isoformat(),
                "sessions": backup_data.sessions,
                "analytics": backup_data.analytics,
                "metadata": backup_data.metadata,
            }

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(backup_dict, f, ensure_ascii=False, indent=2)

            file_size = os.path.getsize(file_path)

            return ExportResult(
                success=True,
                file_path=file_path,
                file_size_bytes=file_size,
                format=ExportFormat.JSON,
                message="数据备份成功"
            )

        except Exception as e:
            return ExportResult(
                success=False,
                file_path=None,
                file_size_bytes=0,
                format=ExportFormat.JSON,
                message=f"数据备份失败：{str(e)}"
            )

    def restore_backup(self, file_path: str) -> Dict[str, Any]:
        """
        恢复数据备份

        Args:
            file_path: 备份文件路径

        Returns:
            恢复的数据字典
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                backup_data = json.load(f)

            # 验证版本
            version = backup_data.get("version", "1.0")
            if version != self.DATA_VERSION:
                return {
                    "success": False,
                    "error": f"版本不兼容：当前{self.DATA_VERSION}, 备份{version}",
                }

            return {
                "success": True,
                "user_id": backup_data.get("user_id"),
                "sessions": backup_data.get("sessions", []),
                "analytics": backup_data.get("analytics", {}),
                "backed_up_at": backup_data.get("backed_up_at"),
            }

        except Exception as e:
            return {
                "success": False,
                "error": f"恢复失败：{str(e)}",
            }

    # =========================================================================
    # Comprehensive Report Generation
    # =========================================================================

    def generate_comprehensive_report(self, user_id: str,
                                      sessions: List[Dict[str, Any]],
                                      learning_profile: Optional[Dict[str, Any]] = None,
                                      behavior_report: Optional[Dict[str, Any]] = None,
                                      statistical_report: Optional[Dict[str, Any]] = None,
                                      recommendation_report: Optional[Dict[str, Any]] = None,
                                      dashboard_data: Optional[Dict[str, Any]] = None,
                                      format: ExportFormat = ExportFormat.PDF) -> ExportResult:
        """
        生成综合报告

        Args:
            user_id: 用户 ID
            sessions: 会话数据
            learning_profile: 学习画像
            behavior_report: 行为报告
            statistical_report: 统计报告
            recommendation_report: 推荐报告
            dashboard_data: 仪表盘数据
            format: 导出格式

        Returns:
            ExportResult 对象
        """
        # 整合所有数据
        report_data = {
            "user_id": user_id,
            "generated_at": datetime.utcnow().isoformat(),
            "period_days": 30,
        }

        # 添加仪表盘数据（如果有）
        if dashboard_data:
            report_data.update(dashboard_data)

        # 添加各模块数据
        if learning_profile:
            report_data["learning_profile"] = learning_profile

        if behavior_report:
            report_data["behavior_report"] = behavior_report

        if statistical_report:
            report_data["statistical_report"] = statistical_report

        if recommendation_report:
            report_data["recommendation_report"] = recommendation_report

        # 根据格式导出
        if format == ExportFormat.PDF:
            return self.export_to_pdf(report_data, report_type=ReportType.COMPLETE)
        elif format == ExportFormat.EXCEL:
            return self.export_to_excel(report_data)
        elif format == ExportFormat.JSON:
            return self.export_to_json(report_data)
        else:
            return self.export_to_json(report_data)

    # =========================================================================
    # Utility Methods
    # =========================================================================

    def get_export_history(self, pattern: str = "*.pdf") -> List[str]:
        """获取导出历史"""
        import glob
        search_pattern = os.path.join(self.output_dir, pattern)
        return sorted(glob.glob(search_pattern))

    def cleanup_old_exports(self, days_old: int = 30) -> int:
        """清理旧导出文件"""
        import glob

        cutoff_date = datetime.utcnow() - timedelta(days=days_old)
        count = 0

        for pattern in ["*.pdf", "*.xlsx", "*.json", "*.csv"]:
            search_pattern = os.path.join(self.output_dir, pattern)
            for file_path in glob.glob(search_pattern):
                file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                if file_time < cutoff_date:
                    try:
                        os.remove(file_path)
                        count += 1
                    except:
                        pass

        return count

    def set_output_dir(self, output_dir: str):
        """设置输出目录"""
        self.output_dir = output_dir
        self._ensure_output_dir()


# Import field from dataclasses
from dataclasses import field
